from flask import Flask, request, jsonify, session, send_from_directory, send_file
from flask_cors import CORS
import sqlite3
import hashlib
import os
import secrets
import sys
import io
try:
    from PIL import Image
except ImportError:
    Image = None

sys.stdout.reconfigure(encoding='utf-8')

app = Flask(__name__)
app.secret_key = "stockpro_secret_key_v1" # Persistent key for stable sessions
CORS(app, supports_credentials=True, origins=["http://localhost:3000"])

DB_PATH = os.path.join(os.path.dirname(__file__), 'database.db')
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ─── Roles ───────────────────────────────────────────────────────────────────
# admin   → бүх зүйл хийх боломжтой
# manager → бараа нэмэх, засах, зарлага, орлого оруулах боломжтой (устгах БҮҮ)
# user    → зөвхөн зарлага/орлого оруулах боломжтой (харах + гүйлгээ хийх)

ROLE_LEVELS = {'user': 1, 'manager': 2, 'admin': 3}


def has_role(min_role: str) -> bool:
    return ROLE_LEVELS.get(session.get('role', ''), 0) >= ROLE_LEVELS[min_role]


def login_required():
    return 'user_id' in session


def unauthorized():
    session.clear()
    return jsonify({'error': 'Нэвтрэх эрхгүй байна'}), 401


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'user'
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS brands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            brand TEXT DEFAULT '',
            barcode TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            category TEXT,
            pack_qty INTEGER DEFAULT 0,
            quantity INTEGER DEFAULT 0,
            price REAL DEFAULT 0,
            price_cn REAL DEFAULT 0,
            has_vat INTEGER DEFAULT 0,
            location_id INTEGER,
            location TEXT DEFAULT 'Үндсэн Агуулах',
            image TEXT,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(location_id) REFERENCES locations(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS transaction_bundles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('in','out','fix')),
            total_amount REAL DEFAULT 0,
            note TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS transaction_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bundle_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL DEFAULT 0,
            has_vat INTEGER DEFAULT 0,
            FOREIGN KEY(bundle_id) REFERENCES transaction_bundles(id),
            FOREIGN KEY(product_id) REFERENCES products(id)
        )
    ''')

    # Keep legacy transactions table for now but we will migrate it
    c.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('in','out','fix')),
            quantity INTEGER NOT NULL,
            note TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(product_id) REFERENCES products(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    # Migrate: add new columns to existing products table if missing
    existing_cols = [row[1] for row in c.execute('PRAGMA table_info(products)').fetchall()]
    for col, coltype in [('brand', 'TEXT DEFAULT ""'), ('barcode', 'TEXT DEFAULT ""'), ('unit', 'TEXT DEFAULT ""'), ('price_cn', 'REAL DEFAULT 0'), ('pack_qty', 'INTEGER DEFAULT 0'), ('has_vat', 'INTEGER DEFAULT 0'), ('location', 'TEXT DEFAULT "Үндсэн Агуулах"'), ('location_id', 'INTEGER')]:
        if col not in existing_cols:
            c.execute(f'ALTER TABLE products ADD COLUMN {col} {coltype}')

    # Ensure default location exists
    c.execute('SELECT * FROM locations WHERE name = ?', ('Үндсэн Агуулах',))
    if not c.fetchone():
        c.execute('INSERT INTO locations (name, description) VALUES (?, ?)', ('Үндсэн Агуулах', 'Админаас үүсгэсэн үндсэн агуулах'))

    # Migration for transaction_bundles type constraint
    bundles_sql = c.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='transaction_bundles'").fetchone()
    if bundles_sql and "CHECK(type IN ('in','out'))" in bundles_sql['sql']:
        c.execute('ALTER TABLE transaction_bundles RENAME TO transaction_bundles_old')
        c.execute('''
            CREATE TABLE transaction_bundles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL CHECK(type IN ('in','out','fix')),
                total_amount REAL DEFAULT 0,
                note TEXT,
                created_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(created_by) REFERENCES users(id)
            )
        ''')
        c.execute('INSERT INTO transaction_bundles SELECT * FROM transaction_bundles_old')
        c.execute('DROP TABLE transaction_bundles_old')
    
    default_loc = c.execute('SELECT id FROM locations WHERE name = ?', ('Үндсэн Агуулах',)).fetchone()
    if default_loc:
        c.execute('UPDATE products SET location_id = ? WHERE location_id IS NULL', (default_loc['id'],))

    # Populate brands from existing products
    c.execute('SELECT DISTINCT brand FROM products WHERE brand IS NOT NULL AND brand != ""')
    existing_brands = c.fetchall()
    for b in existing_brands:
        c.execute('INSERT OR IGNORE INTO brands (name) VALUES (?)', (b['brand'],))

    # Data Migration: Migrate legacy transactions to bundles
    legacy_txs = c.execute('SELECT * FROM transactions').fetchall()
    if legacy_txs:
        for tx in legacy_txs:
            # Check if already migrated (this is a simple heuristic)
            # Actually, let's just create a bundle for each legacy transaction to be safe
            p = c.execute('SELECT price FROM products WHERE id = ?', (tx['product_id'],)).fetchone()
            price = p['price'] if p else 0
            total = tx['quantity'] * price
            
            c.execute('INSERT INTO transaction_bundles (id, type, total_amount, note, created_by, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                      (tx['id'], tx['type'], total, tx['note'], tx['created_by'], tx['created_at']))
            c.execute('INSERT INTO transaction_items (bundle_id, product_id, quantity, price, has_vat) VALUES (?, ?, ?, ?, ?)',
                      (tx['id'], tx['product_id'], tx['quantity'], price, 0))
        
        # Clear legacy transactions to avoid double migration next time
        c.execute('DELETE FROM transactions')

    # Migrate categories from existing products to new table
    existing_cats = c.execute('SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != ""').fetchall()
    for row in existing_cats:
        c.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (row['category'],))

    # Default admin
    c.execute('SELECT * FROM users WHERE username = ?', ('admin',))
    if not c.fetchone():
        c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                  ('admin', hash_password('admin' + '123'), 'admin'))

    conn.commit()
    conn.close()


# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Нэвтрэх нэр болон нууц үг оруулна уу'}), 400
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE username = ? AND password = ?',
                        (username, hash_password(password))).fetchone()
    conn.close()
    if user:
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({'message': 'Амжилттай нэвтэрлээ', 'username': user['username'], 'role': user['role']})
    session.clear()
    return jsonify({'error': 'Нэвтрэх нэр эсвэл нууц үг буруу байна'}), 401


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'message': 'Гарлаа'})


@app.route('/api/me', methods=['GET'])
def me():
    if not login_required():
        return unauthorized()
    return jsonify({'username': session['username'], 'role': session['role']})


@app.route('/api/change-password', methods=['POST'])
def change_password():
    if not login_required():
        return unauthorized()
    data = request.get_json()
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '')
    if not old_pw or not new_pw:
        return jsonify({'error': 'Мэдээлэл дутуу'}), 400
    if len(new_pw) < 6:
        return jsonify({'error': 'Нууц үг хамгийн багадаа 6 тэмдэгтэй байна'}), 400
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ? AND password = ?',
                        (session['user_id'], hash_password(old_pw))).fetchone()
    if not user:
        conn.close()
        session.clear()
        return jsonify({'error': 'Хуучин нууц үг буруу байна'}), 401
    conn.execute('UPDATE users SET password = ? WHERE id = ?',
                 (hash_password(new_pw), session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Нууц үг амжилттай солигдлоо'})


# ─── Users (admin only) ───────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
def get_users():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    conn = get_db()
    users = conn.execute('SELECT id, username, role FROM users ORDER BY id').fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])


@app.route('/api/users', methods=['POST'])
def add_user():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')
    if role not in ('user', 'admin'):
        return jsonify({'error': 'Буруу эрх'}), 400
    if not username or not password:
        return jsonify({'error': 'Мэдээлэл дутуу байна'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Нууц үг хамгийн багадаа 6 тэмдэгт'}), 400
    conn = get_db()
    try:
        conn.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                     (username, hash_password(password), role))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэвтрэх нэр аль хэдийн байна'}), 409
    conn.close()
    return jsonify({'message': 'Хэрэглэгч нэмэгдлээ'}), 201


@app.route('/api/users/<int:uid>', methods=['DELETE'])
def delete_user(uid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    if uid == session['user_id']:
        return jsonify({'error': 'Өөрийгөө устгах боломжгүй'}), 400
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id = ?', (uid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Устгагдлаа'})


@app.route('/api/users/<int:uid>/role', methods=['PUT'])
def change_user_role(uid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    if uid == session['user_id']:
        return jsonify({'error': 'Өөрийн эрхийг өөрчлөх боломжгүй'}), 400
    data = request.get_json()
    role = data.get('role', '')
    if role not in ('user', 'admin'):
        return jsonify({'error': 'Буруу эрх'}), 400
    conn = get_db()
    conn.execute('UPDATE users SET role = ? WHERE id = ?', (role, uid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Эрх өөрчлөгдлөө'})


# ─── Products ─────────────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
def get_products():
    if not login_required():
        return unauthorized()
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    location_id = request.args.get('location_id')
    
    conn = get_db()
    query = '''
        SELECT p.*, l.name as location_name 
        FROM products p
        LEFT JOIN locations l ON p.location_id = l.id
        WHERE 1=1
    '''
    params = []
    if search:
        query += ' AND (p.name LIKE ? OR p.description LIKE ? OR p.brand LIKE ? OR p.barcode LIKE ?)'
        params += [f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%']
    if category:
        query += ' AND p.category = ?'
        params.append(category)
    if location_id:
        query += ' AND p.location_id = ?'
        params.append(location_id)
        
    query += ' ORDER BY p.name ASC'
    products = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(p) for p in products])


@app.route('/api/products', methods=['POST'])
def add_product():
    if not login_required() or not has_role('admin'):
        role = session.get('role', 'none')
        return jsonify({'error': f'Бараа нэмэх эрх байхгүй (Таны эрх: {role})'}), 403
    
    # Handle multipart/form-data if file is present
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.form
        image_file = request.files.get('image')
    else:
        data = request.get_json()
        image_file = None

    name = data.get('name', '').strip()
    brand = data.get('brand', '').strip()
    barcode = data.get('barcode', '').strip()
    unit = data.get('unit', '').strip()
    category = data.get('category', '').strip()
    location_id = data.get('location_id')
    description = data.get('description', '').strip()
    
    # Strict Validation
    if not name:
        return jsonify({'error': 'Барааны нэр заавал оруулна уу'}), 400
    if not unit:
        return jsonify({'error': 'Хэмжих нэгж заавал оруулна уу'}), 400
    if not location_id:
        return jsonify({'error': 'Агуулах заавал оруулна уу'}), 400
    
    try:
        pack_qty = int(data.get('pack_qty', 0))
        quantity = int(data.get('quantity', 0))
        price = float(data.get('price', 0))
        price_cn = float(data.get('price_cn', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Тоон мэдээлэл буруу байна'}), 400
        
    if pack_qty < 0: return jsonify({'error': 'Багцын тоо 0-ээс бага байж болохгүй'}), 400
    if quantity < 0: return jsonify({'error': 'Үлдэгдэл 0-ээс бага байж болохгүй'}), 400
    if pack_qty > 0 and quantity < pack_qty:
        return jsonify({'error': f'Үлдэгдэл нь багцын тооноос бага байж болохгүй (Багц: {pack_qty})'}), 400
    if price < 0: return jsonify({'error': 'Нэгж үнэ 0-ээс бага байж болохгүй'}), 400

    has_vat = 1 if data.get('has_vat') == 'true' or data.get('has_vat') == 1 or data.get('has_vat') is True else 0
    
    img_name = None
    if image_file and image_file.filename:
        ext = image_file.filename.split('.')[-1].lower()
        if ext in ['png', 'jpg', 'jpeg']:
            img_name = f"{secrets.token_hex(8)}.{ext}"
            image_file.save(os.path.join(UPLOAD_FOLDER, img_name))

    conn = get_db()
    if not location_id:
        loc = conn.execute('SELECT id FROM locations LIMIT 1').fetchone()
        location_id = loc['id'] if loc else None

    conn.execute('''
        INSERT INTO products (name, brand, barcode, unit, category, pack_qty, quantity, price, price_cn, has_vat, location_id, description, image) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (name, brand, barcode, unit, category, pack_qty, quantity, price, price_cn, has_vat, location_id, description, img_name))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Бараа нэмэгдлээ'}), 201


@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    if not login_required() or not has_role('admin'):
        role = session.get('role', 'none')
        return jsonify({'error': f'Бараа засах эрх байхгүй (Таны эрх: {role})'}), 403
    
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.form
        image_file = request.files.get('image')
    else:
        data = request.get_json()
        image_file = None

    name = data.get('name', '').strip()
    brand = data.get('brand', '').strip()
    barcode = data.get('barcode', '').strip()
    unit = data.get('unit', '').strip()
    category = data.get('category', '').strip()
    location_id = data.get('location_id')
    description = data.get('description', '').strip()

    # Strict Validation
    if not name:
        return jsonify({'error': 'Барааны нэр заавал оруулна уу'}), 400
    if not unit:
        return jsonify({'error': 'Хэмжих нэгж заавал оруулна уу'}), 400
    if not location_id:
        return jsonify({'error': 'Агуулах заавал оруулна уу'}), 400
    
    try:
        pack_qty = int(data.get('pack_qty', 0))
        quantity = int(data.get('quantity', 0))
        price = float(data.get('price', 0))
        price_cn = float(data.get('price_cn', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Тоон мэдээлэл буруу байна'}), 400

    if pack_qty < 0: return jsonify({'error': 'Багцын тоо 0-ээс бага байж болохгүй'}), 400
    if quantity < 0: return jsonify({'error': 'Үлдэгдэл 0-ээс бага байж болохгүй'}), 400
    if pack_qty > 0 and quantity < pack_qty:
        return jsonify({'error': f'Үлдэгдэл нь багцын тооноос бага байж болохгүй (Багц: {pack_qty})'}), 400
    if price < 0: return jsonify({'error': 'Нэгж үнэ 0-ээс бага байж болохгүй'}), 400

    has_vat = 1 if data.get('has_vat') == 'true' or data.get('has_vat') == 1 or data.get('has_vat') is True else 0
    
    conn = get_db()
    existing = conn.execute('SELECT quantity, image FROM products WHERE id = ?', (pid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Бараа олдсонгүй'}), 404
        
    old_qty = existing['quantity']
    img_name = existing['image']
    
    if image_file and image_file.filename:
        ext = image_file.filename.split('.')[-1].lower()
        if ext in ['png', 'jpg', 'jpeg']:
            if img_name:
                old_path = os.path.join(UPLOAD_FOLDER, img_name)
                if os.path.exists(old_path):
                    try: os.remove(old_path)
                    except: pass
            img_name = f"{secrets.token_hex(8)}.{ext}"
            image_file.save(os.path.join(UPLOAD_FOLDER, img_name))

    # Log quantity change as a 'fix' transaction
    if quantity != old_qty:
        diff = quantity - old_qty
        cursor = conn.execute('INSERT INTO transaction_bundles (type, total_amount, note, created_by) VALUES (?, ?, ?, ?)',
                             ('fix', 0, 'Барааны мэдээлэл засварласнаар үлдэгдэл өөрчлөгдлөө', session['user_id']))
        bundle_id = cursor.lastrowid
        conn.execute('INSERT INTO transaction_items (bundle_id, product_id, quantity, price, has_vat) VALUES (?, ?, ?, ?, ?)',
                     (bundle_id, pid, diff, 0, 0))

    conn.execute('''
        UPDATE products 
        SET name=?, brand=?, barcode=?, unit=?, category=?, pack_qty=?, quantity=?, price=?, price_cn=?, has_vat=?, location_id=?, description=?, image=? 
        WHERE id=?
    ''', (name, brand, barcode, unit, category, pack_qty, quantity, price, price_cn, has_vat, location_id, description, img_name, pid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Бараа шинэчлэгдлээ'})


@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Бараа устгах эрх байхгүй (admin шаардлагатай)'}), 403

    data = request.get_json()
    password = data.get('password')
    if not password:
        return jsonify({'error': 'Нууц үг заавал оруулна уу'}), 400

    conn = get_db()
    user = conn.execute('SELECT password FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user or user['password'] != hash_password(password):
        conn.close()
        return jsonify({'error': 'Нууц үг буруу байна'}), 401

    conn.execute('DELETE FROM products WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Бараа устгагдлаа'})




@app.route('/api/categories', methods=['GET'])
def get_categories_db():
    if not login_required():
        return unauthorized()
    conn = get_db()
    cats = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(c) for c in cats])


@app.route('/api/categories', methods=['POST'])
def add_category():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    name = data.get('name', '').strip()
    desc = data.get('description', '').strip()
    if not name:
        return jsonify({'error': 'Нэр оруулна уу'}), 400
    conn = get_db()
    try:
        conn.execute('INSERT INTO categories (name) VALUES (?)', (name,))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэртэй ангилал аль хэдийн байна'}), 409
    conn.close()
    return jsonify({'message': 'Ангилал нэмэгдлээ'}), 201


@app.route('/api/categories/<int:cid>', methods=['PUT'])
def update_category(cid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    name = data.get('name', '').strip()
    desc = data.get('description', '').strip()
    if not name:
        return jsonify({'error': 'Нэр оруулна уу'}), 400
    conn = get_db()
    try:
        conn.execute('UPDATE categories SET name=? WHERE id=?', (name, cid))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэртэй ангилал аль хэдийн байна'}), 409
    conn.close()
    return jsonify({'message': 'Ангилал шинэчлэгдлээ'})


@app.route('/api/categories/<int:cid>', methods=['DELETE'])
def delete_category(cid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй (admin шаардлагатай)'}), 403
    conn = get_db()
    # Optional: check if any products use this category name?
    # For now, just delete the category from the list. 
    # Products store category as a string, so they won't break but might lose the reference in the list.
    conn.execute('DELETE FROM categories WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Ангилал устгагдлаа'})


@app.route('/api/locations', methods=['GET'])
def get_locations():
    if not login_required():
        return unauthorized()
    conn = get_db()
    locs = conn.execute("SELECT * FROM locations ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(l) for l in locs])


@app.route('/api/locations', methods=['POST'])
def add_location():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    name = data.get('name', '').strip()
    desc = data.get('description', '').strip()
    if not name:
        return jsonify({'error': 'Нэр оруулна уу'}), 400
    conn = get_db()
    try:
        conn.execute('INSERT INTO locations (name, description) VALUES (?, ?)', (name, desc))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэртэй агуулах аль хэдийн байна'}), 409
    conn.close()
    return jsonify({'message': 'Агуулах нэмэгдлээ'}), 201


@app.route('/api/locations/<int:lid>', methods=['DELETE'])
def delete_location(lid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй (admin шаардлагатай)'}), 403
    conn = get_db()
    # Check if any products are in this location
    count = conn.execute('SELECT COUNT(*) FROM products WHERE location_id = ?', (lid,)).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({'error': f'Энэ агуулахад {count} бараа байгаа тул устгах боломжгүй'}), 400
    
    conn.execute('DELETE FROM locations WHERE id = ?', (lid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Агуулах устгагдлаа'})


# ─── Transactions (зарлага / орлого) ─────────────────────────────────────────

@app.route('/api/transactions/bundle', methods=['POST'])
def add_transaction_bundle():
    if not login_required():
        return unauthorized()
    data = request.get_json()
    items = data.get('items', [])
    tx_type = data.get('type')  # 'in' or 'out'
    note = data.get('note', '').strip()

    if not items or tx_type not in ('in', 'out'):
        return jsonify({'error': 'Мэдээлэл дутуу байна'}), 400

    conn = get_db()
    total_amount = 0
    
    try:
        # Validate all items first and calculate total
        for item in items:
            p_id = item.get('product_id')
            qty = int(item.get('quantity', 0))
            if qty <= 0: continue
            
            product = conn.execute('SELECT quantity, name, price FROM products WHERE id = ?', (p_id,)).fetchone()
            if not product:
                raise ValueError(f'Бараа олдсонгүй (ID: {p_id})')
            
            if tx_type == 'out' and product['quantity'] < qty:
                raise ValueError(f"'{product['name']}' барааны үлдэгдэл хүрэлцэхгүй байна (Үлдэгдэл: {product['quantity']})")
            
            # For "in" transactions (Add Stock), always use DB price. For "out" (Sell), use frontend price.
            price = product['price'] if tx_type == 'in' else float(item.get('price', 0))
            
            # Pack-aware Stock Validation
            db_pack_qty = product['pack_qty'] or 1
            if tx_type == 'out' and qty > product['quantity']:
                raise ValueError(f"'{product['name']}' барааны үлдэгдэл хүрэлцэхгүй байна (Үлдэгдэл: {product['quantity']} ш)")
            
            # (Packs are derived, so we don't need to validate a separate 'packs' input anymore)
            total_amount += qty * price

        # Create Bundle
        cursor = conn.execute('INSERT INTO transaction_bundles (type, total_amount, note, created_by) VALUES (?, ?, ?, ?)',
                             (tx_type, total_amount, note, session['user_id']))
        bundle_id = cursor.lastrowid

        # Process Items
        for item in items:
            p_id = item.get('product_id')
            qty = int(item.get('quantity', 0))
            if qty <= 0: continue
            
            # Fetch price again or use from validated set
            p = conn.execute('SELECT price FROM products WHERE id = ?', (p_id,)).fetchone()
            # For "in" (Add Stock), use DB price. For "out" (Sell), use frontend price.
            price = p['price'] if tx_type == 'in' else float(item.get('price', 0))
            
            has_vat = 1 if item.get('has_vat') else 0
            
            delta = qty if tx_type == 'in' else -qty
            conn.execute('UPDATE products SET quantity = quantity + ? WHERE id = ?', (delta, p_id))
            conn.execute('INSERT INTO transaction_items (bundle_id, product_id, quantity, price, has_vat) VALUES (?, ?, ?, ?, ?)',
                         (bundle_id, p_id, qty, price, has_vat))
            
        conn.commit()
        conn.close()
        action = 'Орлого' if tx_type == 'in' else 'Зарлага'
        return jsonify({'message': f'{action} амжилттай бүртгэгдлээ', 'bundle_id': bundle_id}), 201
    except Exception as e:
        if 'conn' in locals(): conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/brands', methods=['GET'])
def get_brands():
    if not login_required():
        return unauthorized()
    conn = get_db()
    brands = conn.execute('SELECT * FROM brands ORDER BY name ASC').fetchall()
    conn.close()
    return jsonify([dict(b) for b in brands])

@app.route('/api/brands', methods=['POST'])
def add_brand():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Брэндийн нэр хоосон байна'}), 400
    
    conn = get_db()
    try:
        conn.execute('INSERT INTO brands (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Брэнд амжилттай үүсгэгдлээ'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэртэй брэнд аль хэдийн байна'}), 409
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/<int:bid>', methods=['PUT'])
def update_brand(bid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй'}), 403
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Нэр хоосон байна'}), 400
    
    conn = get_db()
    try:
        conn.execute('UPDATE brands SET name=? WHERE id=?', (name, bid))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Брэнд амжилттай шинэчлэгдлээ'})
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ийм нэртэй брэнд аль хэдийн байна'}), 409
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/<int:bid>', methods=['DELETE'])
def delete_brand(bid):
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Зөвшөөрөл байхгүй (admin шаардлагатай)'}), 403
    conn = get_db()
    try:
        conn.execute('DELETE FROM brands WHERE id=?', (bid,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Брэнд устгагдлаа'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    if not login_required():
        return unauthorized()
    tx_type = request.args.get('type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    limit = min(int(request.args.get('limit', 100)), 500)

    conn = get_db()
    query = '''
        SELECT b.*, u.username as created_by 
        FROM transaction_bundles b
        JOIN users u ON b.created_by = u.id
        WHERE 1=1
    '''
    params = []
    if tx_type in ('in', 'out'):
        query += ' AND b.type = ?'
        params.append(tx_type)
    if start_date:
        query += ' AND DATE(b.created_at) >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND DATE(b.created_at) <= ?'
        params.append(end_date)
    query += ' ORDER BY b.created_at DESC LIMIT ?'
    params.append(limit)
    
    bundles = conn.execute(query, params).fetchall()
    result = []
    for b in bundles:
        bundle_dict = dict(b)
        items = conn.execute('''
            SELECT ti.*, p.name as product_name, p.brand, p.barcode, p.pack_qty, p.unit
            FROM transaction_items ti
            JOIN products p ON ti.product_id = p.id
            WHERE ti.bundle_id = ?
        ''', (b['id'],)).fetchall()
        bundle_dict['items'] = [dict(i) for i in items]
        result.append(bundle_dict)
        
    conn.close()
    return jsonify(result)


# ─── Stats ────────────────────────────────────────────────────────────────────

@app.route('/api/stats', methods=['GET'])
def get_stats():
    if not login_required():
        return unauthorized()
    conn = get_db()
    total_products = conn.execute('SELECT COUNT(*) as c FROM products').fetchone()['c']
    total_qty = conn.execute('SELECT COALESCE(SUM(quantity),0) as s FROM products').fetchone()['s']
    total_value = conn.execute('SELECT COALESCE(SUM(quantity*price),0) as v FROM products').fetchone()['v']
    # Low stock: quantity < 20% of pack_qty (if pack_qty > 0) or quantity <= 5
    low_stock = conn.execute("SELECT COUNT(*) as c FROM products WHERE (pack_qty > 0 AND quantity < 0.2 * pack_qty) OR (pack_qty = 0 AND quantity <= 5)").fetchone()['c']
    
    # Revenue/Expense stats
    today_out = conn.execute("SELECT COALESCE(SUM(total_amount),0) as s FROM transaction_bundles WHERE type='out' AND date(created_at)=date('now')").fetchone()['s']
    today_in = conn.execute("SELECT COALESCE(SUM(total_amount),0) as s FROM transaction_bundles WHERE type='in' AND date(created_at)=date('now')").fetchone()['s']
    
    week_out = conn.execute("SELECT COALESCE(SUM(total_amount),0) as s FROM transaction_bundles WHERE type='out' AND date(created_at) >= date('now', '-7 days')").fetchone()['s']
    month_out = conn.execute("SELECT COALESCE(SUM(total_amount),0) as s FROM transaction_bundles WHERE type='out' AND date(created_at) >= date('now', '-30 days')").fetchone()['s']

    conn.close()
    return jsonify({
        'total_products': total_products,
        'total_quantity': total_qty,
        'total_value': round(total_value, 2),
        'low_stock': low_stock,
        'today_out_amount': today_out,
        'today_in_amount': today_in,
        'week_out_amount': week_out,
        'month_out_amount': month_out
    })


@app.route('/api/stats/revenue', methods=['GET'])
def get_stats_revenue():
    if not login_required():
        return unauthorized()
    
    period = request.args.get('period', 'monthly')
    conn = get_db()
    
    if period == 'annually':
        interval = '-1 year'
        group_by = "strftime('%Y-%m', created_at)"
    else:  # monthly (default)
        interval = '-30 days'
        group_by = "strftime('%Y-%m-%d', created_at)"

    query = f'''
        SELECT {group_by} as day,
               SUM(CASE WHEN type='out' THEN total_amount ELSE 0 END) as income,
               SUM(CASE WHEN type='in' THEN total_amount ELSE 0 END) as expense
        FROM transaction_bundles
        WHERE created_at >= date('now', '{interval}')
        GROUP BY day
        ORDER BY day ASC
    '''
    rows = conn.execute(query).fetchall()
    
    # Distribution by Location
    loc_dist = conn.execute('''
        SELECT l.name, SUM(p.quantity * p.price) as value
        FROM products p
        JOIN locations l ON p.location_id = l.id
        GROUP BY l.name
    ''').fetchall()
    
    # Distribution by Category
    cat_dist = conn.execute('''
        SELECT category, SUM(quantity * price) as value
        FROM products p
        WHERE category != ''
        GROUP BY category
    ''').fetchall()
    
    conn.close()
    return jsonify({
        'revenue_chart': [dict(r) for r in rows],
        'location_distribution': [dict(l) for l in loc_dist],
        'category_distribution': [dict(c) for c in cat_dist]
    })

@app.route('/api/stats/product-trend', methods=['GET'])
def get_stats_product_trend():
    if not login_required():
        return unauthorized()
    
    period = request.args.get('period', 'monthly')
    conn = get_db()
    
    if period == 'weekly':
        interval = '-7 days'
        group_by = "strftime('%Y-%m-%d', created_at)"
    elif period == 'annually':
        interval = '-1 year'
        group_by = "strftime('%Y-%m', created_at)"
    else:  # monthly
        interval = '-30 days'
        group_by = "strftime('%Y-%m-%d', created_at)"

    query = f'''
        SELECT {group_by} as date, COUNT(*) as count
        FROM products
        WHERE created_at >= date('now', '{interval}')
        GROUP BY date
        ORDER BY date ASC
    '''
    rows = conn.execute(query).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ─── Excel Import ─────────────────────────────────────────────────────────────

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def parse_excel(file_bytes):
    """Excel файл уншиж, мөр бүрийг dict болгон буцаана."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = None
    col_map = {}

    FIELD_ALIASES = {
        'name':     ['бараа нэр', 'нэр'],
        'unit':     ['нэгж'],
        'qty_new':  ['тоо ширхэг', 'тоо'],
        'qty_rem':  ['үлдэгдэл'],
        'price_cn': ['урдаас ирсэн үнэ юань', 'юань', 'үнэ юань'],
        'price':    ['төгрөг', 'үнэ төгрөг', 'мнт', 'үнийн дүн'],
        'brand':    ['брэнд'],
        'category': ['ангилал', 'төрөл'],
        'code':     ['бараа код', 'код', 'баркод'],
        'image_col':['зураг'],
        'location': ['агуулах'],
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell is None:
                continue
            val = str(cell).strip().lower()
            for field, aliases in FIELD_ALIASES.items():
                if any(a in val for a in aliases):
                    if field not in col_map:
                        col_map[field] = col_idx
        if col_map.get('name'):
            header_row = row_idx
            break

    if not header_row or not col_map.get('name'):
        return None, '"Бараа нэр" гарчиг олдсонгүй. Excel файлаа шалгана уу.'

    # Extract images
    image_map = {} # (row, col) -> filename
    if hasattr(ws, '_images'):
        for img in ws._images:
            # openpyxl stores image anchors. Usually OneCellAnchor or TwoCellAnchor.
            # We want to know which row it's in.
            row = None
            col = None
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
            if hasattr(img.anchor, '_from'):
                row = img.anchor._from.row + 1 # 1-indexed
                col = img.anchor._from.col + 1
            elif hasattr(img.anchor, 'row'):
                row = img.anchor.row
                col = img.anchor.col
            
            if row is not None and Image:
                img_name = f"import_{secrets.token_hex(8)}.png"
                img_path = os.path.join(UPLOAD_FOLDER, img_name)
                # Save image using Pillow
                try:
                    pil_img = Image.open(io.BytesIO(img._data()))
                    pil_img.save(img_path)
                    image_map[(row, col)] = img_name
                except Exception as e:
                    print(f"Image extraction error: {e}")

    rows = []
    empty_row_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        def g(field):
            idx = col_map.get(field)
            return row[idx - 1].value if idx and idx <= len(row) else None

        name = g('name')
        if not name or str(name).strip() == '':
            empty_row_count += 1
            if empty_row_count >= 10: # Stop after 10 empty rows for performance
                break
            continue
        
        empty_row_count = 0 # Reset if name is found
        name = str(name).strip()

        def clean_num(val):
            if val is None: return 0
            s = str(val).replace('₮', '').replace('¥', '').replace(',', '').strip()
            try:
                return float(s)
            except:
                return 0

        try:
            qty_new = int(clean_num(g('qty_new')))
        except:
            qty_new = 0
        try:
            qty_rem = int(clean_num(g('qty_rem')))
        except:
            qty_rem = qty_new
        try:
            price = clean_num(g('price'))
        except:
            price = 0.0
        try:
            price_cn = clean_num(g('price_cn'))
        except:
            price_cn = 0.0

        brand = str(g('brand') or '').strip()
        category = str(g('category') or '').strip()
        unit = str(g('unit') or '').strip()
        code = g('code')
        code = str(code) if code is not None else ''
        
        loc_val = str(g('location') or '').strip() or 'Үндсэн Агуулах'
        
        # Check for image in this row at image_col
        img_filename = None
        img_col_idx = col_map.get('image_col')
        if img_col_idx:
            pass

        rows.append({
            'name': name,
            'brand': brand,
            'category': category,
            'unit': unit,
            'code': code,
            'qty_new': qty_new,
            'qty_rem': qty_rem,
            'price': price,
            'price_cn': clean_num(g('price_cn')),
            'location': loc_val,
            'image_file': None # Will fill below
        })

    # Since values_only=True doesn't give us row index easily, 
    # let's re-iterate with index to match images.
    for i, r_data in enumerate(rows):
        row_num = header_row + 1 + i
        img_col_idx = col_map.get('image_col')
        if img_col_idx and (row_num, img_col_idx) in image_map:
            r_data['image_file'] = image_map[(row_num, img_col_idx)]

    return rows, None


@app.route('/api/import/products', methods=['POST'])
def import_products_excel():
    """Excel-ээс бараа импортлох (manager+)"""
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Эрх байхгүй'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'Файл олдсонгүй'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'error': 'Зөвхөн .xlsx файл зөвшөөрнө'}), 400
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl суулгаагүй байна'}), 500

    file_bytes = f.read()
    rows, err = parse_excel(file_bytes)
    if err:
        return jsonify({'error': err}), 400

    mode = request.form.get('mode', 'skip')
    added = 0
    updated = 0
    skipped = 0
    errors = []

    conn = get_db()
    for r in rows:
        name = r['name']
        brand = r['brand'] or ''
        barcode = r['code'] or ''
        unit = r['unit'] or ''
        category = r['category'] or brand or 'Бусад'
        location_name = r.get('location', 'Үндсэн Агуулах') or 'Үндсэн Агуулах'
        pack_qty_excel = r['qty_new']
        qty_excel = r['qty_rem'] if r['qty_rem'] > 0 else r['qty_new']
        price = r['price']
        
        # Validation
        if pack_qty_excel > 0 and qty_excel < pack_qty_excel:
            errors.append(f"'{name}' - Үлдэгдэл нь багцын тооноос бага байж болохгүй")
            skipped += 1
            continue
            
        image_file = r.get('image_file')

        # Resolve location_id from name
        loc_row = conn.execute('SELECT id FROM locations WHERE name = ?', (location_name,)).fetchone()
        if not loc_row:
            # If location doesn't exist, create it to be robust
            conn.execute('INSERT INTO locations (name) VALUES (?)', (location_name,))
            loc_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        else:
            loc_id = loc_row['id']

        # Register Brand if exists and not present
        if brand:
            conn.execute('INSERT OR IGNORE INTO brands (name) VALUES (?)', (brand,))
            
        # Register Category if exists and not present
        if category:
            conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (category,))

        # Check for existing product by Name AND Barcode
        existing = conn.execute('SELECT id, pack_qty, quantity, image FROM products WHERE name = ? AND barcode = ?', (name, barcode)).fetchone()
        
        try:
            if existing:
                # 1. Match Found: Increment stock and update location
                new_pack_qty = existing['pack_qty'] + pack_qty_excel
                new_qty = existing['quantity'] + qty_excel
                
                # Use new image if provided, otherwise keep existing
                final_img = image_file if image_file else existing['image']
                
                conn.execute('''
                    UPDATE products 
                    SET pack_qty=?, quantity=?, location_id=?, location=?, image=? 
                    WHERE id=?
                ''', (new_pack_qty, new_qty, loc_id, location_name, final_img, existing['id']))
                updated += 1
            else:
                # 2. New Product: Create entry
                conn.execute('''
                    INSERT INTO products (name, brand, barcode, unit, category, pack_qty, quantity, price, price_cn, location_id, location, image) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, brand, barcode, unit, category, pack_qty_excel, qty_excel, price, r.get('price_cn', 0), loc_id, location_name, image_file))
                added += 1
        except Exception as ex:
            errors.append(f'{name}: {ex}')

    conn.commit()
    conn.close()

    return jsonify({
        'message': f'{added} шинэ бараа бүртгэгдлээ, {updated} барааны үлдэгдэл нэмэгдлээ',
        'added': added, 'updated': updated, 'skipped': skipped,
        'errors': errors[:10]
    })


@app.route('/api/import/transactions', methods=['POST'])
def import_transactions_excel():
    """Excel-ээс гүйлгээ (орлого/зарлага) импортлох"""
    if not login_required():
        return unauthorized()
    if 'file' not in request.files:
        return jsonify({'error': 'Файл олдсонгүй'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'error': 'Зөвхөн .xlsx файл зөвшөөрнө'}), 400
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl суулгаагүй байна'}), 500

    tx_type = request.form.get('type', 'in')
    if tx_type not in ('in', 'out'):
        return jsonify({'error': 'Төрөл буруу (in эсвэл out)'}), 400

    note = request.form.get('note', '').strip()

    file_bytes = f.read()
    rows, err = parse_excel(file_bytes)
    if err:
        return jsonify({'error': err}), 400

    added = 0
    skipped = 0
    not_found = []
    errors = []

    conn = get_db()
    for r in rows:
        name = r['name']
        qty = r['qty_new'] if tx_type == 'in' else r['qty_rem']
        if qty <= 0:
            qty = r['qty_new'] or r['qty_rem']
        if qty <= 0:
            skipped += 1
            continue

        product = conn.execute('SELECT * FROM products WHERE name = ?', (name,)).fetchone()
        if not product:
            not_found.append(name)
            skipped += 1
            continue

        if tx_type == 'out' and product['quantity'] < qty:
            errors.append(f'{name}: үлдэгдэл хүрэлцэхгүй ({product["quantity"]} < {qty})')
            skipped += 1
            continue

        try:
            delta = qty if tx_type == 'in' else -qty
            conn.execute('UPDATE products SET quantity = quantity + ? WHERE id = ?', (delta, product['id']))
            conn.execute('INSERT INTO transactions (product_id, type, quantity, note, created_by) VALUES (?,?,?,?,?)',
                         (product['id'], tx_type, qty, note or 'Excel импорт', session['user_id']))
            added += 1
        except Exception as ex:
            errors.append(f'{name}: {ex}')
            skipped += 1

    conn.commit()
    conn.close()

    action = 'Орлого' if tx_type == 'in' else 'Зарлага'
    msg = f'{action}: {added} бараа бүртгэгдлээ'
    if skipped:
        msg += f', {skipped} алгасагдлаа'

    return jsonify({
        'message': msg,
        'added': added, 'skipped': skipped,
        'not_found': not_found[:20],
        'errors': errors[:10]
    })


# ─── Excel Export ─────────────────────────────────────────────────────────────

@app.route('/api/export/products', methods=['GET'])
def export_products():
    """Бүх барааг Excel файлаар татах"""
    if not login_required():
        return unauthorized()
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl суулгаагүй байна'}), 500

    conn = get_db()
    products = conn.execute('SELECT * FROM products ORDER BY name ASC').fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Барааны жагсаалт'
    
    # New strict order: Брэнд, Бараа код, Зураг, Бараа нэр, Нэгж, Тоо Ширхэг, Үлдэгдэл, Урдаас ирсэн үнэ Юань, Төгрөг, Агуулах
    headers = ['Брэнд', 'Бараа код', 'Зураг', 'Бараа нэр', 'Нэгж', 'Тоо Ширхэг', 'Үлдэгдэл', 'Урдаас ирсэн үнэ Юань', 'Төгрөг', 'Агуулах']
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    from openpyxl.drawing.image import Image as XLImage
    
    for idx, p in enumerate(products, 2): # Start from row 2
        img_val = ''
        if p['image']:
            img_path = os.path.join(UPLOAD_FOLDER, p['image'])
            if os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    # Resize for cell (roughly 50x50px)
                    img.width = 40
                    img.height = 40
                    # Zuraag 3-r baganad (Column C) oruulna
                    ws.add_image(img, f'C{idx}')
                    ws.row_dimensions[idx].height = 35 # Adjust row height
                except Exception as e:
                    img_val = f'Error: {e}'
        
        ws.append([
            p['brand'] or '',
            p['barcode'] or '',
            img_val, # Cell text can be empty if image is added
            p['name'],
            p['unit'] or '',
            p['pack_qty'] or 0,
            p['quantity'],
            f"{p['price_cn'] or 0} ¥",
            f"{p['price'] or 0} ₮",
            p['location'] or 'Үндсэн Агуулах'
        ])

    # Apply borders and alignment
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.row > 1: # Data rows
                cell.alignment = Alignment(vertical='center', horizontal='center')

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='products_export.xlsx'
    )


# ─── Template Download ────────────────────────────────────────────────────────

@app.route('/api/template', methods=['GET'])
def download_template():
    """Import template файл татах"""
    if os.path.exists(TEMPLATE_PATH):
        return send_file(TEMPLATE_PATH, as_attachment=True, download_name='template.xlsx')
    # Generate template on the fly
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl суулгаагүй байна'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Бараа импорт'

    from openpyxl.styles import Font, PatternFill, Alignment
    # New strict order: Брэнд, Бараа код, Зураг, Бараа нэр, Нэгж, Тоо Ширхэг, Үлдэгдэл, Урдаас ирсэн үнэ Юань, Төгрөг, Агуулах
    headers = ['Брэнд', 'Бараа код', 'Зураг', 'Бараа нэр', 'Нэгж', 'Тоо Ширхэг', 'Үлдэгдэл', 'Урдаас ирсэн үнэ Юань', 'Төгрөг', 'Агуулах']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    ws.append(['Samsung', '001', '', 'Galaxy S21', 'ширхэг', 10, 10, '500 ¥', '500,000 ₮', 'Үндсэн агуулах'])
    ws.append(['LG', '002', '', 'LG Monitor 27', 'хайрцаг', 5, 5, '800 ¥', '1,200,000 ₮', 'Үндсэн агуулах'])

    # Apply borders
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', horizontal='center')

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template.xlsx')


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


if __name__ == '__main__':
    init_db()
    # Migration: add image column if not exists
    conn = get_db()
    try:
        conn.execute('ALTER TABLE products ADD COLUMN image TEXT')
        conn.commit()
    except:
        pass
    conn.close()
    
    print("Server started: http://localhost:5000")
    # Admin login info removed for security
    print("Roles: admin > manager > user")
    app.run(debug=True, port=5000)
