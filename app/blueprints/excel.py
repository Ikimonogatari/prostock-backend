from flask import Blueprint, request, jsonify, send_file, current_app, session
from app.database import get_db
from app.utils import login_required, unauthorized, has_role
import os
import io
import secrets
import secrets
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

excel_bp = Blueprint('excel', __name__)

def parse_excel(file_bytes):
    """Excel файл уншиж, мөр бүрийг dict болгон буцаана."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = None
    col_map = {}

    FIELD_ALIASES = {
        'name':     ['бараа нэр', 'нэр'],
        'unit':     ['нэгж'],
        'qty_new':  ['тоо ширхэг', 'тоо'],
        'qty_rem':  ['үлдэгдэл'],
        'price_cn': ['урдаас ирсэн үнэ юань', 'юань', 'үнэ юань'],
        'price':    ['төгрөг', 'үнэ төгрөг', 'мнт', 'үнийн дүн'],
        'brand':    ['брэнд'],
        'category': ['ангилал', 'төрөл'],
        'code':     ['бараа код', 'код', 'баркод'],
        'image_col':['зураг'],
        'location': ['агуулах'],
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell is None:
                continue
            val = str(cell).strip().lower()
            for field, aliases in FIELD_ALIASES.items():
                if any(a in val for a in aliases):
                    if field not in col_map:
                        col_map[field] = col_idx
        if col_map.get('name'):
            header_row = row_idx
            break

    if not header_row or not col_map.get('name'):
        return None, '"Бараа нэр" гарчиг олдсонгүй. Excel файлаа шалгана уу.'

    # Extract images
    image_map = {} # (row, col) -> filename
    if hasattr(ws, '_images'):
        for img in ws._images:
            row = None
            col = None
            if hasattr(img.anchor, '_from'):
                row = img.anchor._from.row + 1 # 1-indexed
                col = img.anchor._from.col + 1
            elif hasattr(img.anchor, 'row'):
                row = img.anchor.row
                col = img.anchor.col
            
            if row is not None and HAS_PILLOW:
                img_name = f"import_{secrets.token_hex(8)}.png"
                img_path = os.path.join(current_app.config['UPLOAD_FOLDER'], img_name)
                try:
                    pil_img = Image.open(io.BytesIO(img._data()))
                    pil_img.save(img_path)
                    image_map[(row, col)] = img_name
                except Exception as e:
                    print(f"Image extraction error: {e}")

    rows = []
    empty_row_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        def g(field):
            idx = col_map.get(field)
            return row[idx - 1].value if idx and idx <= len(row) else None

        name = g('name')
        if not name or str(name).strip() == '':
            empty_row_count += 1
            if empty_row_count >= 10:
                break
            continue
        
        empty_row_count = 0
        name = str(name).strip()

        def clean_num(val):
            if val is None: return 0
            s = str(val).replace('₮', '').replace('¥', '').replace(',', '').strip()
            try:
                return float(s)
            except:
                return 0

        qty_new = int(clean_num(g('qty_new')))
        qty_rem = int(clean_num(g('qty_rem'))) or qty_new
        price = clean_num(g('price'))
        price_cn = clean_num(g('price_cn'))
        brand = str(g('brand') or '').strip()
        category = str(g('category') or '').strip()
        unit = str(g('unit') or '').strip()
        code = str(g('code') or '').strip()
        loc_val = str(g('location') or '').strip() or 'Үндсэн Агуулах'
        
        rows.append({
            'name': name,
            'brand': brand,
            'category': category,
            'unit': unit,
            'code': code,
            'qty_new': qty_new,
            'qty_rem': qty_rem,
            'price': price,
            'price_cn': price_cn,
            'location': loc_val,
            'image_file': None
        })

    for i, r_data in enumerate(rows):
        row_num = header_row + 1 + i
        img_col_idx = col_map.get('image_col')
        if img_col_idx and (row_num, img_col_idx) in image_map:
            r_data['image_file'] = image_map[(row_num, img_col_idx)]

    return rows, None

@excel_bp.route('/import/products', methods=['POST'])
def import_products():
    if not login_required() or not has_role('admin'):
        return jsonify({'error': 'Эрх байхгүй'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'Файл олдсонгүй'}), 400
    f = request.files['file']
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl суулгаагүй байна'}), 500

    file_bytes = f.read()
    mode = request.form.get('mode', 'update')
    form_location_id = request.form.get('location_id')
    if form_location_id == 'all':
        return jsonify({'error': 'Импорт хийхийн тулд тодорхой агуулах сонгоно уу.'}), 400
    
    rows, err = parse_excel(file_bytes)
    if err: return jsonify({'error': err}), 400

    added, updated, skipped = 0, 0, 0
    errors = []
    conn = get_db()
    
    # Create transaction bundle for import
    bundle_id = None
    try:
        cur = conn.execute('INSERT INTO transaction_bundles (type, total_amount, note, created_by) VALUES (?, ?, ?, ?)',
                           ('in', 0, 'Excel импортоор барааны үлдэгдэл өөрчилсөн', session.get('user_id')))
        bundle_id = cur.lastrowid
    except:
        pass
        
    total_bundle_amount = 0
    for r in rows:
        try:
            name, brand, barcode = r['name'], r['brand'], r['code']
            category = r['category'] or brand or 'Бусад'
            location_name = r['location']
            pack_qty_excel, qty_excel = r['qty_new'], r['qty_rem']
            price, price_cn = r['price'], r['price_cn']
            image_file = r['image_file']

            # Location
            if form_location_id:
                loc_id = int(form_location_id)
                loc_row = conn.execute('SELECT name FROM locations WHERE id = ?', (loc_id,)).fetchone()
                location_name = loc_row['name'] if loc_row else 'Үндсэн Агуулах'
            else:
                loc = conn.execute('SELECT id FROM locations WHERE name = ?', (location_name,)).fetchone()
                if not loc:
                    conn.execute('INSERT INTO locations (name) VALUES (?)', (location_name,))
                    loc_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                else: 
                    loc_id = loc['id']

            # Metadata
            if brand: conn.execute('INSERT OR IGNORE INTO brands (name) VALUES (?)', (brand,))
            if category: conn.execute('INSERT OR IGNORE INTO categories (name) VALUES (?)', (category,))

            # 1. Check if the exact product exists at this local location
            # Note: Barcode might be empty, so we check (barcode OR name) AND location_id
            existing_local = conn.execute('''
                SELECT id, pack_qty, quantity, image 
                FROM products 
                WHERE location_id = ? AND ((barcode != '' AND barcode = ?) OR name = ?)
            ''', (loc_id, barcode, name)).fetchone()

            if existing_local:
                new_pack = existing_local['pack_qty'] + pack_qty_excel
                new_qty = existing_local['quantity'] + qty_excel
                final_img = image_file or existing_local['image']
                conn.execute('UPDATE products SET pack_qty=?, quantity=?, image=? WHERE id=?',
                             (new_pack, new_qty, final_img, existing_local['id']))
                
                if bundle_id and qty_excel != 0:
                    conn.execute('INSERT INTO transaction_items (bundle_id, product_id, quantity, price, has_vat) VALUES (?, ?, ?, ?, ?)',
                                 (bundle_id, existing_local['id'], qty_excel, price, 0))
                    if qty_excel > 0:
                        total_bundle_amount += qty_excel * price
                updated += 1
            else:
                # 2. It doesn't exist locally. But does it exist globally? Let's borrow its stats if so.
                existing_global = conn.execute('''
                    SELECT name, brand, barcode, unit, category, description, image, has_vat
                    FROM products 
                    WHERE ((barcode != '' AND barcode = ?) OR name = ?)
                    LIMIT 1
                ''', (barcode, name)).fetchone()

                if existing_global:
                    final_name = existing_global['name']
                    final_brand = existing_global['brand']
                    final_barcode = existing_global['barcode']
                    final_unit = existing_global['unit']
                    final_category = existing_global['category']
                    final_image = image_file or existing_global['image']
                    final_vat = existing_global['has_vat']
                else:
                    final_name = name
                    final_brand = brand
                    final_barcode = barcode
                    final_unit = r.get('unit', '')
                    final_category = category
                    final_image = image_file
                    final_vat = 0

                cursor = conn.execute('''
                    INSERT INTO products (name, brand, barcode, unit, category, pack_qty, quantity, price, price_cn, has_vat, location_id, location, image, description) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (final_name, final_brand, final_barcode, final_unit, final_category, pack_qty_excel, qty_excel, price, price_cn, final_vat, loc_id, location_name, final_image, r.get('description', '')))
                new_id = cursor.lastrowid
                
                # Global sync: If we actually introduced new data (like an image), let's sync it back globally
                if final_image or final_brand or final_category:
                     sync_query = '''
                         UPDATE products 
                         SET brand=?, category=?, image=?, has_vat=? 
                         WHERE id != ? AND ((barcode=? AND barcode!='') OR (name=? AND name!=''))
                     '''
                     conn.execute(sync_query, (final_brand, final_category, final_image, final_vat, new_id, final_barcode, final_name))
                
                if bundle_id and qty_excel != 0:
                    conn.execute('INSERT INTO transaction_items (bundle_id, product_id, quantity, price, has_vat) VALUES (?, ?, ?, ?, ?)',
                                 (bundle_id, new_id, qty_excel, price, 0))
                    if qty_excel > 0:
                        total_bundle_amount += qty_excel * price
                added += 1
        except Exception as ex: errors.append(f'{r["name"]}: {ex}')

    if bundle_id:
        conn.execute('UPDATE transaction_bundles SET total_amount = ? WHERE id = ?', (total_bundle_amount, bundle_id))

    conn.commit(); conn.close()
    return jsonify({'message': f'{added} нэмэгдсэн, {updated} шинэчлэгдсэн', 'added': added, 'updated': updated, 'errors': errors[:10]})

@excel_bp.route('/export/products', methods=['GET'])
def export_products():
    if not login_required(): return unauthorized()
    if not HAS_OPENPYXL: return jsonify({'error': 'openpyxl суулгаагүй'}), 500

    conn = get_db()
    location_id = request.args.get('location_id', '')
    if location_id:
        products = conn.execute('SELECT * FROM products WHERE location_id = ? ORDER BY name ASC', (location_id,)).fetchall()
    else:
        products = conn.execute('SELECT * FROM products ORDER BY name ASC').fetchall()
    conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Барааны жагсаалт'
    headers = ['Брэнд', 'Ангилал', 'Бараа код', 'Зураг', 'Бараа нэр', 'Нэгж', 'Тоо Ширхэг', 'Үлдэгдэл', 'Урдаас ирсэн үнэ Юань', 'Төгрөг', 'Агуулах']
    ws.append(headers)

    # Styling
    header_font = Font(bold=True, color='FFFFFF'); header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')

    for idx, p in enumerate(products, 2):
        img_val = ''
        if p['image']:
            img_path = os.path.join(current_app.config['UPLOAD_FOLDER'], p['image'])
            if os.path.exists(img_path):
                try:
                    img = XLImage(img_path); img.width = img.height = 40
                    ws.add_image(img, f'C{idx}'); ws.row_dimensions[idx].height = 35
                except: img_val = 'Error'
        
        ws.append([p['brand'] or '', p['category'] or '', p['barcode'] or '', img_val, p['name'], p['unit'] or '', p['pack_qty'] or 0, p['quantity'], f"{p['price_cn'] or 0} ¥", f"{p['price'] or 0} ₮", p['location'] or 'Үндсэн Агуулах'])

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='products_export.xlsx')

@excel_bp.route('/template', methods=['GET'])
def download_template():
    if not HAS_OPENPYXL: return jsonify({'error': 'openpyxl суулгаагүй'}), 500
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Бараа импорт'
    headers = ['Брэнд', 'Ангилал', 'Бараа код', 'Зураг', 'Бараа нэр', 'Нэгж', 'Тоо Ширхэг', 'Үлдэгдэл', 'Урдаас ирсэн үнэ Юань', 'Төгрөг', 'Агуулах']
    ws.append(headers)
    ws.append(['Samsung', 'Утас', '001', '', 'Galaxy S21', 'ширхэг', 10, 10, '500 ¥', '500,000 ₮', 'Үндсэн агуулах'])
    
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='template.xlsx')
